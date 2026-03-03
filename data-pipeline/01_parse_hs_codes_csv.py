#!/usr/bin/env python3
"""
Script 01: Parse HS Code CSV/XLSX from data.go.th → cg_hs_codes

No AI needed — pure data parsing.
Uses ON CONFLICT (upsert) for idempotency.

Provenance: source_url = https://data.go.th/en/dataset/hscode
"""

import os
import sys
import csv
import json
import glob

sys.path.insert(0, os.path.dirname(__file__))
from config import RAW_DIR, get_db_conn, PROVENANCE_SOURCES
from utils.state_tracker import StateTracker
from utils.validator import validate_hs_code, normalize_hs_code, validate_rate

SOURCE_URL = PROVENANCE_SOURCES["data_go_th"]
INPUT_DIR = os.path.join(RAW_DIR, "hs-codes")


def parse_csv_file(filepath: str) -> list[dict]:
    """Parse a CSV file into HS code records."""
    records = []
    encoding_options = ["utf-8", "utf-8-sig", "tis-620", "cp874"]

    for encoding in encoding_options:
        try:
            with open(filepath, "r", encoding=encoding) as f:
                # Try to detect delimiter
                sample = f.read(4096)
                f.seek(0)

                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                reader = csv.DictReader(f, dialect=dialect)

                for row in reader:
                    record = extract_hs_record(row)
                    if record:
                        records.append(record)
                break  # Success with this encoding
        except (UnicodeDecodeError, csv.Error):
            continue
        except Exception as e:
            print(f"  Error with {encoding}: {e}")
            continue

    return records


def parse_xlsx_file(filepath: str) -> list[dict]:
    """Parse an XLSX file into HS code records."""
    import openpyxl

    records = []
    wb = openpyxl.load_workbook(filepath, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First row as headers
        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]

        for row in rows[1:]:
            row_dict = dict(zip(headers, row))
            record = extract_hs_record(row_dict)
            if record:
                records.append(record)

    wb.close()
    return records


def extract_hs_record(row: dict) -> dict | None:
    """
    Extract an HS code record from a row dict.
    Handles various column name conventions from Thai government data.
    """
    # Map common column names (Thai + English variants)
    code = None
    for key in ["code", "hs_code", "hscode", "พิกัด", "รหัส", "tariff_code",
                 "hs code", "tariffcode", "hscodeth"]:
        if key in row and row[key]:
            code = str(row[key]).strip()
            break

    if not code:
        # Try first column that looks like an HS code
        for key, val in row.items():
            if val and isinstance(val, (str, int, float)):
                val_str = str(val).strip().replace(".", "")
                if val_str.isdigit() and 4 <= len(val_str) <= 12:
                    code = str(val).strip()
                    break

    if not code:
        return None

    # Validate
    valid, msg = validate_hs_code(code)
    if not valid:
        return None

    code = normalize_hs_code(code)

    # Extract description
    desc_th = None
    for key in ["description_th", "รายละเอียด", "คำอธิบาย", "รายการ",
                 "description_thai", "th_desc", "desc_th"]:
        if key in row and row[key]:
            desc_th = str(row[key]).strip()
            break

    desc_en = None
    for key in ["description_en", "description", "description_eng",
                 "en_desc", "desc_en", "eng_description"]:
        if key in row and row[key]:
            desc_en = str(row[key]).strip()
            break

    # Extract rate
    base_rate = None
    for key in ["rate", "base_rate", "duty_rate", "อัตราอากร", "อัตรา",
                 "tariff_rate", "mfn_rate"]:
        if key in row and row[key] is not None:
            try:
                rate_str = str(row[key]).strip().replace("%", "")
                base_rate = float(rate_str)
                valid_rate, _ = validate_rate(base_rate)
                if not valid_rate:
                    base_rate = None
            except (ValueError, TypeError):
                pass
            break

    # Extract other fields
    unit = None
    for key in ["unit", "หน่วย"]:
        if key in row and row[key]:
            unit = str(row[key]).strip()
            break

    category = None
    for key in ["category", "หมวด", "ประเภท"]:
        if key in row and row[key]:
            category = str(row[key]).strip()
            break

    # Parse section/chapter/heading from code
    digits = code.replace(".", "")
    chapter = int(digits[:2]) if len(digits) >= 2 else None
    heading = code[:7] if len(code) >= 7 else None
    subheading = code[:10] if len(code) >= 10 else None

    return {
        "code": code,
        "section": None,  # Would need a section lookup table
        "chapter": chapter,
        "heading": heading,
        "subheading": subheading,
        "description_th": desc_th,
        "description_en": desc_en,
        "base_rate": base_rate,
        "unit": unit,
        "category": category,
        "source_url": SOURCE_URL,
    }


def upsert_hs_codes(conn, records: list[dict]):
    """Insert or update HS codes using ON CONFLICT."""
    inserted = 0
    updated = 0

    with conn.cursor() as cur:
        for rec in records:
            cur.execute("""
                INSERT INTO cg_hs_codes
                    (code, section, chapter, heading, subheading,
                     description_th, description_en, base_rate, unit, category,
                     updated_at)
                VALUES
                    (%(code)s, %(section)s, %(chapter)s, %(heading)s, %(subheading)s,
                     %(description_th)s, %(description_en)s, %(base_rate)s, %(unit)s, %(category)s,
                     NOW())
                ON CONFLICT (code) DO UPDATE SET
                    description_th = COALESCE(EXCLUDED.description_th, cg_hs_codes.description_th),
                    description_en = COALESCE(EXCLUDED.description_en, cg_hs_codes.description_en),
                    base_rate = COALESCE(EXCLUDED.base_rate, cg_hs_codes.base_rate),
                    unit = COALESCE(EXCLUDED.unit, cg_hs_codes.unit),
                    category = COALESCE(EXCLUDED.category, cg_hs_codes.category),
                    chapter = COALESCE(EXCLUDED.chapter, cg_hs_codes.chapter),
                    heading = COALESCE(EXCLUDED.heading, cg_hs_codes.heading),
                    subheading = COALESCE(EXCLUDED.subheading, cg_hs_codes.subheading),
                    updated_at = NOW()
            """, rec)

            if cur.statusmessage.startswith("INSERT"):
                inserted += 1
            else:
                updated += 1

    conn.commit()
    return inserted, updated


def main():
    if not os.path.exists(INPUT_DIR):
        print(f"ERROR: Input directory not found: {INPUT_DIR}")
        print("Run 00_collect_raw_data.py first!")
        sys.exit(1)

    conn = get_db_conn()
    tracker = StateTracker(conn, "01_parse_hs_codes_csv")

    # Find all CSV/XLSX files
    files = sorted(
        glob.glob(os.path.join(INPUT_DIR, "*.csv"))
        + glob.glob(os.path.join(INPUT_DIR, "*.xlsx"))
    )

    if not files:
        print(f"No CSV/XLSX files found in {INPUT_DIR}")
        sys.exit(1)

    print(f"Found {len(files)} data files")
    total_inserted = 0
    total_updated = 0

    for filepath in files:
        filename = os.path.basename(filepath)
        source_key = f"csv:{filename}"

        if tracker.is_processed(source_key):
            print(f"  Skip (already done): {filename}")
            continue

        print(f"\nParsing: {filename}")
        try:
            if filename.endswith(".xlsx"):
                records = parse_xlsx_file(filepath)
            else:
                records = parse_csv_file(filepath)

            print(f"  Extracted {len(records)} HS code records")

            if records:
                inserted, updated = upsert_hs_codes(conn, records)
                total_inserted += inserted
                total_updated += updated
                print(f"  DB: {inserted} inserted, {updated} updated")
                tracker.mark_processed(source_key, f"{inserted} inserted, {updated} updated")
            else:
                tracker.mark_skipped(source_key, "no valid records found")

        except Exception as e:
            print(f"  ERROR: {e}")
            tracker.mark_failed(source_key, str(e))

    # Final count
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM cg_hs_codes")
        total = cur.fetchone()[0]

    print(f"\n{'='*40}")
    print(f"Total HS codes in DB: {total}")
    print(f"This run: {total_inserted} inserted, {total_updated} updated")
    print(f"Progress: {tracker.get_progress()}")

    conn.close()


if __name__ == "__main__":
    main()
